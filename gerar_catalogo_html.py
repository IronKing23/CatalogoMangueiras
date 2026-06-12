# -*- coding: utf-8 -*-
"""
GERADOR DE CATÁLOGO HTML INTERATIVO — Mangueiras Colhedora John Deere CH570
Usina Cedro · Copecar
============================================================================
Lê a aba CATÁLOGO do Excel + a pasta de diagramas e gera um catálogo HTML
auto-contido (funciona offline) com:

 NAVEGAÇÃO   • cards recolhíveis (+ abrir/fechar todos) • link direto por
               sistema (#ST______) • atalho "/" para buscar, Esc para limpar
               • botão topo • no celular: tabela em linhas com rolagem lateral
 FUNÇÕES     • lista de requisição (carrinho) com quantidades, PDF de
               requisição e exportação CSV • "onde essa peça é usada" ao
               clicar no OEM • ordenação clicando nos cabeçalhos • copiar OEM
               • chips de filtro por diâmetro / família / proteção
               • destaque do termo buscado • PDF estruturado dos itens filtrados
 VISUAL      • bolinha de cor por diâmetro (legenda do catálogo) • logos
               automáticos (logo_cedro.png e logo_grupo.png ao lado do script) • zoom com
               arrastar/rolar/pinça nos diagramas • cabeçalho de tabela fixo

USO:
    python gerar_catalogo_html.py
    python gerar_catalogo_html.py --excel MeuCatalogo.xlsx --diagramas diagramas --saida catalogo.html

FLUXO: edite o Excel (linhas, colunas novas, observações) e rode o script —
colunas novas entram automaticamente. REQUISITOS: pip install openpyxl
"""
import argparse, base64, html, json, re, sys, unicodedata
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Instale a dependência: pip install openpyxl")

# ----------------------------- CONFIGURAÇÃO ---------------------------------
ABA_CATALOGO = "CATÁLOGO"
COL_SISTEMA, COL_ST = "SISTEMA", "ST"
COLUNAS_IGNORADAS = {"Nº", "DIAGRAMA"}
ABREVIACOES = {
    "MANGUEIRA AEROQUIP": "MANGUEIRA",
    "COMPRIMENTO CORTE (m)": "COMPR. (m)",
    "CONEXÃO 1 AEROQUIP": "CONEXÃO 1",
    "CONEXÃO 2 AEROQUIP": "CONEXÃO 2",
    "PROTEÇÃO AEROQUIP": "PROTEÇÃO",
    "OBSERVAÇÕES": "OBS.",
}
COL_MANGUEIRA = "MANGUEIRA AEROQUIP"   # origem do diâmetro/família
COL_PROTECAO = "PROTEÇÃO AEROQUIP"     # origem do chip de proteção
LOGOS_CEDRO = ["logo_cedro.png", "logo_cedro.jpg", "logo.png", "logo.jpg",
               "diagramas/logo_cedro.png", "diagramas/logo.png"]
LOGOS_GRUPO = ["logo_grupo.png", "logo_grupo.jpg", "logo_pedra.png", "logo_pedra.jpg",
               "diagramas/logo_grupo.png", "diagramas/logo_pedra.png"]
ENDERECO_UNIDADE = ("Usina Cedro · Rod. BR-158, Km 62, Zona Rural, Paranaíba/MS · "
                    "(67) 3669-7405 · Pedra Agroindustrial")

CORES_DIAM = {4: "#2244cc", 6: "#e6c317", 8: "#222222", 10: "#7a3fb0",
              12: "#cc2222", 20: "#b8860b", 24: "#e64ba0", 32: "#e64ba0"}
ROTULOS_DIAM = {4: 'Ø1/4"', 6: 'Ø3/8"', 8: 'Ø1/2"', 10: 'Ø5/8"', 12: 'Ø3/4"',
                16: 'Ø1"', 20: 'Ø1.1/4"', 24: 'Ø1.1/2"', 32: 'Ø2"'}

# ----------------------------- AUXILIARES -----------------------------------
def normaliza(txt):
    txt = unicodedata.normalize("NFD", str(txt))
    return "".join(c for c in txt if unicodedata.category(c) != "Mn").lower()

def fmt(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:g}".replace(".", ",")
    return str(v).strip()

def b64(caminho):
    ext = caminho.suffix.lower().lstrip(".")
    mime = "jpeg" if ext in ("jpg", "jpeg") else ext
    return f"data:image/{mime};base64," + base64.b64encode(caminho.read_bytes()).decode()

def parse_mangueira(valor):
    """'GH781-8 / FC735-O8' -> diâmetros {8} e famílias {'GH781','FC735'}."""
    diams, fams = set(), set()
    for seg in valor.split("/"):
        seg = seg.strip()
        if not seg:
            continue
        m = re.search(r"-([0O]?\d+)\s*$", seg)
        if m:
            diams.add(int(m.group(1).replace("O", "0")))
            fam = re.sub(r"-[0O]?\d+\s*$", "", seg).strip()
        else:
            fam = seg
        if fam:
            fams.add(fam)
    return diams, fams

def ponto_diam(d):
    rot = ROTULOS_DIAM.get(d, f"-{d}")
    if d == 16:
        estilo = "background:conic-gradient(#2e8b3a 0 120deg,#e07b1f 0 240deg,#fff 0 360deg)"
        rot += " — cor varia por tramas (ver legenda)"
    elif d in CORES_DIAM:
        estilo = f"background:{CORES_DIAM[d]}"
    else:
        estilo = "background:#999"
    return f'<span class="pt" style="{estilo}" title="{html.escape(rot)}"></span>'

def le_excel(arquivo):
    wb = load_workbook(arquivo, data_only=True)
    if ABA_CATALOGO not in wb.sheetnames:
        sys.exit(f"Aba '{ABA_CATALOGO}' não encontrada em {arquivo}")
    ws = wb[ABA_CATALOGO]
    linhas = ws.iter_rows(values_only=True)
    cabecalho = [fmt(c) for c in next(linhas)]
    sistemas, ordem = {}, []
    for row in linhas:
        reg = dict(zip(cabecalho, [fmt(v) for v in row]))
        if not any(reg.values()):
            continue
        chave = (reg.get(COL_SISTEMA, ""), reg.get(COL_ST, ""))
        if chave not in sistemas:
            sistemas[chave] = []
            ordem.append(chave)
        sistemas[chave].append(reg)
    colunas = [c for c in cabecalho if c and c not in COLUNAS_IGNORADAS
               and c not in (COL_SISTEMA, COL_ST)]
    return ordem, sistemas, colunas

# ----------------------------- TEMPLATE --------------------------------------
TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Catálogo de Mangueiras — John Deere CH570 · Usina Cedro</title>
<style>
:root { --verde-escuro:#3E5F33; --verde:#6B9A57; --verde-suave:#8FB97C;
        --verde-pastel:#E7F1DF; --palha:#E4D8AE; --creme:#F8F8F1;
        --tinta:#2E3B28; --topo-h:120px; }
* { box-sizing:border-box; }
html { scroll-behavior:smooth; }
body { margin:0; font-family:Calibri,'Segoe UI',Arial,sans-serif;
       background:var(--creme); color:var(--tinta); }

/* ===== Barra superior ===== */
.topo { position:sticky; top:0; z-index:60; color:#fff; padding:12px 18px;
        background:linear-gradient(135deg,var(--verde-suave),var(--verde) 55%,var(--verde-escuro));
        box-shadow:0 2px 10px rgba(46,59,40,.30); }
.topo-linha1 { display:flex; align-items:center; gap:12px; }
.logo { height:48px; filter:drop-shadow(0 1px 3px rgba(0,0,0,.35)); }
.logo-dir { margin-left:auto; }
@media (max-width:760px){ .logo-dir { display:none; } }
.topo h1 { margin:0; font-size:1.15rem; line-height:1.25; }
.topo .sub { font-size:.78rem; opacity:.92; }
.controles { display:flex; flex-wrap:wrap; gap:8px; margin-top:9px; align-items:center; }
.controles input[type=search], .controles select { padding:8px 12px; border:none;
    border-radius:6px; font-size:.95rem; background:#FDFDF8; color:var(--tinta); }
.controles input[type=search] { flex:1 1 230px; }
.controles select { max-width:330px; }
.controles button { padding:7px 12px; border:1px solid #FDFDF8; background:transparent;
                    color:#fff; border-radius:6px; cursor:pointer; font-size:.88rem; }
.controles button:hover { background:rgba(255,255,255,.18); }
.controles button.destaque { background:var(--palha); border-color:var(--palha);
                             color:var(--verde-escuro); font-weight:bold; }
.controles button.destaque:hover { filter:brightness(1.05); }
#contador { font-size:.84rem; white-space:nowrap; }
#painel-filtros { display:none; margin-top:10px; background:rgba(255,255,255,.13);
                  border-radius:8px; padding:10px 12px; }
#painel-filtros .grupo { display:flex; flex-wrap:wrap; gap:6px; align-items:center; margin:5px 0; }
#painel-filtros .grupo b { font-size:.78rem; min-width:74px; font-weight:600; opacity:.95; }
.chip { padding:3px 10px; border-radius:14px; border:1px solid #fff; background:transparent;
        color:#fff; cursor:pointer; font-size:.78rem; display:inline-flex; gap:5px; align-items:center; }
.chip.on { background:#fff; color:var(--verde-escuro); font-weight:bold; }
.chip .pt { width:11px; height:11px; }
#limpar-chips { font-size:.76rem; text-decoration:underline; background:none; border:none;
                color:#fff; cursor:pointer; }

/* ===== Estado vazio ===== */
#vazio { display:none; text-align:center; color:#6F7E63; padding:64px 12px; }
#vazio p { font-size:1.05rem; margin-bottom:14px; }
#vazio button { padding:9px 20px; border-radius:6px; border:1px solid var(--verde);
                background:#fff; color:var(--verde-escuro); cursor:pointer; font-size:.9rem; }

/* ===== Cards ===== */
main { margin:18px auto 60px; max-width:1240px; padding:0 16px; }
.card { background:#fff; border:1px solid #D9E5CE; border-radius:10px; margin-bottom:14px;
        overflow:hidden; box-shadow:0 1px 4px rgba(62,95,51,.10);
        scroll-margin-top:calc(var(--topo-h) + 10px); }
.card-cab { display:flex; align-items:center; gap:10px; flex-wrap:wrap; cursor:pointer;
            background:var(--verde-pastel); border-bottom:2px solid var(--verde); padding:9px 14px; }
.card-cab h2 { margin:0; font-size:.97rem; color:var(--verde-escuro); flex:1 1 55%; }
.badge { background:var(--verde); color:#fff; padding:2px 10px; border-radius:12px;
         font-size:.76rem; font-weight:bold; }
.qtd { font-size:.76rem; color:#6F7E63; }
.ic-link { border:none; background:transparent; cursor:pointer; font-size:.95rem; opacity:.55; }
.ic-link:hover { opacity:1; }
.seta { font-size:.8rem; color:var(--verde-escuro); transition:transform .15s; }
.card.aberto .seta { transform:rotate(90deg); }
.card-corpo { display:none; padding:12px 14px; }
.card.aberto .card-corpo { display:block; }
.bloco-img { text-align:center; margin-bottom:10px; }
.bloco-img img { max-height:300px; max-width:100%; border:1px solid #E2EAD9; border-radius:6px;
                 cursor:zoom-in; margin:0 6px 6px 0; background:#fff; }
.sem-img { color:#9AA68F; font-size:.84rem; margin:0; }

table { border-collapse:collapse; width:100%; table-layout:fixed; font-size:.85rem; }
th { background:var(--verde); color:#fff; padding:6px 4px; text-align:center; font-size:.78rem;
     overflow-wrap:break-word; position:sticky; top:calc(var(--topo-h) - 1px); z-index:5; }
th.th-sort { cursor:pointer; }
th.th-sort:hover { background:var(--verde-escuro); }
th .dir { font-size:.65rem; }
th.c-sel, td.c-sel { width:30px; }
th.c-dia, td.c-dia { width:58px; }
td { border:1px solid #E6EDDE; padding:5px 4px; text-align:center; overflow-wrap:break-word; }
tbody tr:nth-child(even) { background:#F2F7EC; }
tbody tr.flash { animation:flash 2.2s; }
@keyframes flash { 0%,60% { background:var(--palha); } 100% { background:inherit; } }
tr.oculta, section.oculta { display:none; }
.pt { display:inline-block; width:14px; height:14px; border-radius:50%;
      border:1px solid rgba(0,0,0,.3); vertical-align:-2px; }
.dnum { font-size:.72rem; color:#666; margin-left:3px; }
mark { background:var(--palha); padding:0 1px; border-radius:2px; }
.c-oem .v { cursor:pointer; text-decoration:underline dotted; }
.c-oem .v:hover { color:var(--verde-escuro); font-weight:bold; }
.ic-copia { border:none; background:transparent; cursor:pointer; font-size:.85rem;
            opacity:.45; margin-left:3px; }
tr:hover .ic-copia { opacity:1; }

/* ===== Flutuantes / gaveta / modais ===== */
.flutuante { position:fixed; right:18px; width:46px; height:46px; border-radius:50%;
             border:none; background:var(--verde); color:#fff; font-size:1.15rem; cursor:pointer;
             box-shadow:0 3px 10px rgba(0,0,0,.3); z-index:70; }
.flutuante:hover { background:var(--verde-escuro); }
#btn-cesta { bottom:18px; }
#btn-topo { bottom:76px; display:none; }
#badge-cesta { position:absolute; top:-6px; right:-6px; background:#c0392b; color:#fff;
               border-radius:10px; font-size:.68rem; padding:1px 6px; font-weight:bold; }
#cesta { position:fixed; right:0; top:0; bottom:0; width:400px; max-width:96vw; background:#fff;
         z-index:90; box-shadow:-4px 0 18px rgba(0,0,0,.3); transform:translateX(105%);
         transition:transform .2s; display:flex; flex-direction:column; }
#cesta.aberta { transform:none; }
.cesta-cab { background:var(--verde); color:#fff; padding:12px 16px; font-weight:bold;
             display:flex; justify-content:space-between; align-items:center; }
.cesta-cab button { background:none; border:none; color:#fff; font-size:1.1rem; cursor:pointer; }
#cesta-corpo { flex:1; overflow-y:auto; padding:0 12px 10px; }
#cesta-corpo table { font-size:.78rem; }
#cesta-corpo th { position:sticky; top:0; z-index:2; box-shadow:0 2px 4px rgba(0,0,0,.15); }
#cesta-corpo td { padding:4px 3px; }
#cesta-corpo input[type=number] { width:46px; padding:2px; }
#cesta-corpo .rm { border:none; background:none; color:#c0392b; cursor:pointer; font-weight:bold; }
.cesta-pe { padding:10px 12px; border-top:1px solid #E2EAD9; display:flex; gap:8px; flex-wrap:wrap; }
.cesta-pe button { flex:1 1 45%; padding:8px; border-radius:6px; border:1px solid var(--verde);
                   background:#fff; color:var(--verde-escuro); cursor:pointer; font-size:.84rem; }
.cesta-pe button.ok { background:var(--verde); color:#fff; font-weight:bold; }
#aviso { position:fixed; bottom:24px; left:50%; transform:translateX(-50%);
         background:var(--verde-escuro); color:#fff; padding:9px 18px; border-radius:20px;
         font-size:.85rem; z-index:120; display:none; box-shadow:0 3px 12px rgba(0,0,0,.35); }

.veu { display:none; position:fixed; inset:0; background:rgba(20,30,16,.65); z-index:80; }
.veu.aberto { display:flex; align-items:center; justify-content:center; }
#modal { z-index:100; background:rgba(20,30,16,.88); overflow:hidden; touch-action:none; }
#modal img { position:absolute; left:0; top:0; transform-origin:0 0; max-width:none;
             border-radius:4px; background:#fff; cursor:grab; user-select:none; }
.modal-botoes { position:absolute; top:12px; right:14px; display:flex; gap:6px; z-index:5; }
.modal-botoes button { width:38px; height:38px; border-radius:8px; border:none; font-size:1.05rem;
                       background:rgba(255,255,255,.92); cursor:pointer; }
.caixa { background:#fff; border-radius:10px; padding:20px 24px; max-width:480px; width:92%;
         border-top:6px solid var(--verde); box-shadow:0 8px 30px rgba(0,0,0,.35);
         max-height:84vh; overflow-y:auto; }
.caixa h3 { margin:0 0 8px; color:var(--verde-escuro); }
.caixa p { margin:6px 0; font-size:.9rem; }
.caixa label { display:flex; gap:8px; align-items:center; margin:11px 0; font-size:.93rem; }
.caixa .acoes { display:flex; gap:10px; justify-content:flex-end; margin-top:13px; }
.caixa .acoes button { padding:8px 16px; border-radius:6px; border:1px solid var(--verde);
                       cursor:pointer; font-size:.9rem; background:#fff; color:var(--verde-escuro); }
.caixa .acoes button.ok { background:var(--verde); color:#fff; font-weight:bold; }
#lista-uso { list-style:none; margin:8px 0 0; padding:0; }
#lista-uso li { padding:7px 4px; border-bottom:1px solid #EEF3E8; font-size:.85rem;
                cursor:pointer; }
#lista-uso li:hover { background:var(--verde-pastel); }
#lista-uso .st { color:#6F7E63; font-size:.75rem; }
footer { text-align:center; color:#9AA68F; font-size:.78rem; padding:24px 0 36px; }

@media screen and (max-width:760px) {
  .topo { padding:10px 12px; }
  .topo h1 { font-size:.98rem; }
  .controles input[type=search] { flex:1 1 100%; }
  .controles select { flex:1 1 100%; max-width:none; }
  .controles button { padding:6px 9px; font-size:.78rem; }
  #contador { flex-basis:100%; }
  .card-cab h2 { font-size:.9rem; }
  .bloco-img img { max-height:220px; }
  /* tabela em linhas, com rolagem lateral suave; cabeçalho, seleção e ordenação visíveis */
  .card-corpo { padding:10px 8px; overflow-x:auto; -webkit-overflow-scrolling:touch; }
  .card-corpo table { min-width:720px; }
  .card-corpo th { position:static; }
}

/* ===== Impressão (relatório / requisição) ===== */
#relatorio { display:none; }
@media print {
  @page { size:A4; margin:13mm 11mm; }
  body.imprimindo > *:not(#relatorio) { display:none !important; }
  body.imprimindo #relatorio { display:block; }
  #relatorio { font-size:10pt; color:#1c2418; -webkit-print-color-adjust:exact; print-color-adjust:exact; }
  #relatorio table { display:table; width:100%; }
  #relatorio thead { display:table-header-group; }
  #relatorio tbody { display:table-row-group; }
  #relatorio tr { display:table-row; }
  #relatorio th, #relatorio td { display:table-cell; }
  .rel-topo { background:var(--verde);
              background:linear-gradient(135deg,var(--verde-suave),var(--verde) 55%,var(--verde-escuro));
              color:#fff; border-radius:3mm; padding:4.5mm 6mm; display:flex;
              align-items:center; gap:5mm; }
  .rel-logo { height:13mm; filter:drop-shadow(0 0.4mm 0.8mm rgba(0,0,0,.35)); }
  .rel-logo.dir { margin-left:4mm; }
  .rel-tit h1 { margin:0; font-size:14pt; color:#fff; }
  .rel-tit .linha2 { font-size:8.8pt; opacity:.93; margin-top:1mm; }
  .rel-doc { background:#fff; color:var(--verde-escuro); font-weight:bold; font-size:10pt;
             padding:2mm 3.5mm; border-radius:2mm; margin-left:auto; white-space:nowrap; }
  .rel-tit { margin-right:auto; }
  .rel-fio { height:1.4mm; background:var(--palha); border-radius:1mm; margin:2mm 0 4mm; }
  .rel-filtros { background:var(--verde-pastel); border-left:1.4mm solid var(--verde);
                 border-radius:0 2mm 2mm 0; padding:2.5mm 3.5mm; font-size:9.5pt;
                 margin-bottom:4mm; }
  .rel-sis { margin-bottom:13px; }
  .rel-sis h3 { background:var(--verde); color:#fff; font-size:10.5pt; padding:5px 8px;
                margin:0; border-radius:3px 3px 0 0; break-after:avoid; break-inside:avoid; }
  .rel-sis img { display:block; max-width:76%; max-height:92mm; margin:6px auto; break-inside:avoid; }
  #relatorio table { width:100%; border-collapse:collapse; table-layout:fixed; font-size:8.5pt; }
  #relatorio th { background:var(--verde); color:#fff; position:static;
                  border:.5pt solid var(--verde-escuro); padding:3.5px 2px; font-size:8pt; }
  #relatorio td { border:.5pt solid #C5D6B8; padding:3px 2px; text-align:center; overflow-wrap:break-word; }
  #relatorio tbody tr:nth-child(even) { background:#F4F9F0; }
  .req-meta { display:flex; gap:4mm; margin-bottom:4mm; }
  .req-meta div { flex:1; background:#fff; border:.5pt solid #C5D6B8; border-radius:2mm;
                  padding:2mm 3mm; font-size:10pt; }
  .req-meta b { color:var(--verde-escuro); display:block; font-size:7.6pt;
                text-transform:uppercase; letter-spacing:.3pt; }
  #relatorio td.qtd { background:var(--palha); font-weight:bold; font-size:9.5pt; }
  .req-total { text-align:right; font-size:10pt; margin-top:2mm; }
  .req-obs { border:.5pt solid #C5D6B8; border-radius:2mm; padding:2mm 3mm;
             margin-top:4mm; font-size:9pt; break-inside:avoid; }
  .req-obs b { color:var(--verde-escuro); font-size:8pt; text-transform:uppercase; }
  .req-obs .linha { border-bottom:.4pt solid #9DB98C; height:6.5mm; }
  .req-assin { border:.5pt solid #C5D6B8; border-radius:2mm; margin-top:4.5mm;
               break-inside:avoid; overflow:hidden; }
  .req-assin h4 { margin:0; background:var(--verde-pastel); color:var(--verde-escuro);
                  font-size:8.5pt; text-transform:uppercase; letter-spacing:.4pt;
                  padding:2mm 3mm; }
  .req-assin .campos { display:flex; gap:10mm; padding:9mm 8mm 4mm; }
  .req-assin .campo { flex:1; text-align:center; font-size:9pt; }
  .req-assin .campo .linha-ass { border-bottom:.6pt solid #333; height:9mm; margin-bottom:1.5mm; }
  .req-assin .campo span { color:#6F7E63; font-size:7.8pt; }
  .rel-rodape { margin-top:14px; font-size:8.5pt; color:#6F7E63; text-align:center; }
}
</style>
</head>
<body>
<div class="topo" id="topo">
  <div class="topo-linha1">
    __LOGO__
    <div>
      <h1>Catálogo de Mangueiras — Colhedora John Deere CH570</h1>
      <div class="sub">Usina Cedro · Pedra Agroindustrial · Catálogo Copecar · __NSIS__ sistemas · __NLIN__ mangueiras · gerado em __DATA__</div>
    </div>
    __LOGOG__
  </div>
  <div class="controles">
    <input id="busca" type="search" placeholder='Buscar OEM, mangueira, conexão, sistema…  (atalho: tecla "/")'>
    <select id="filtro"><option value="">Todos os sistemas</option>__OPCOES__</select>
    <span id="contador"></span>
    <button onclick="togglePainel()">Filtros ▾</button>
    __BTNLEG__
    <button class="destaque" onclick="abrePdfOpcoes()">Gerar PDF</button>
    <button onclick="limpa()">Limpar</button>
    <button onclick="abreTodos(true)" title="Expandir todos os sistemas">▾ todos</button>
    <button onclick="abreTodos(false)" title="Recolher todos os sistemas">▸ todos</button>
  </div>
  <div id="painel-filtros">
    <div class="grupo"><b>Diâmetro</b>__CHIPS_D__</div>
    <div class="grupo"><b>Família</b>__CHIPS_F__</div>
    <div class="grupo"><b>Proteção</b>__CHIPS_P__ <button id="limpar-chips" onclick="limpaChips()">limpar filtros</button></div>
  </div>
</div>

<main id="lista">__CARDS__<div id="vazio">
  <p>🌾 Nenhuma mangueira encontrada com esses filtros.</p>
  <button onclick="limpa()">Limpar filtros</button>
</div></main>

<button class="flutuante" id="btn-topo" title="Voltar ao topo" onclick="window.scrollTo({top:0,behavior:'smooth'})">↑</button>
<button class="flutuante" id="btn-cesta" title="Lista de requisição" onclick="toggleCesta()">🛒<span id="badge-cesta">0</span></button>

<div id="cesta">
  <div class="cesta-cab"><span>Lista de requisição</span><button onclick="toggleCesta()">✕</button></div>
  <div id="cesta-corpo"></div>
  <div class="cesta-pe">
    <button class="ok" onclick="gerarRequisicao()">Requisição em PDF</button>
    <button onclick="exportaCSV()">Exportar CSV</button>
    <button onclick="limpaCesta()">Limpar lista</button>
  </div>
</div>

<div id="modal" class="veu">
  <div class="modal-botoes">
    <button onclick="zoomBtn(1.3)" title="Aproximar">＋</button>
    <button onclick="zoomBtn(0.77)" title="Afastar">－</button>
    <button onclick="zoomReset()" title="Tamanho original">⟳</button>
    <button onclick="fechaModal()" title="Fechar">✕</button>
  </div>
  <img id="modal-img" src="" alt="Zoom">
</div>

<div id="modal-uso" class="veu" onclick="if(event.target===this)this.classList.remove('aberto')">
  <div class="caixa">
    <h3 id="uso-titulo"></h3>
    <p style="color:#6F7E63;font-size:.82rem">Clique para ir ao sistema.</p>
    <ul id="lista-uso"></ul>
    <div class="acoes"><button onclick="document.getElementById('modal-uso').classList.remove('aberto')">Fechar</button></div>
  </div>
</div>

<div id="pdf-opcoes" class="veu">
  <div class="caixa">
    <h3>Gerar PDF das peças filtradas</h3>
    <p id="pdf-resumo"></p>
    <label><input type="checkbox" id="chk-img" checked> Incluir diagramas dos sistemas</label>
    <p style="color:#6F7E63;font-size:.82rem">Na janela de impressão, escolha o destino <b>“Salvar como PDF”</b>.</p>
    <div class="acoes">
      <button onclick="fechaPdfOpcoes()">Cancelar</button>
      <button class="ok" onclick="gerarPDF()">Gerar</button>
    </div>
  </div>
</div>

<div id="relatorio"></div>
<div id="aviso"></div>
<footer>__ENDERECO__<br>Gerado por gerar_catalogo_html.py · edite o Excel e rode o script novamente para atualizar</footer>

<script>
'use strict';
const COLS = __COLSJS__;
const FULLCOLS = __FULLCOLSJS__;
const OEM_IDX = __OEMIDX__;
const legendaSrc = "__LEG__";
const logoCedro = "__LOGOC__";
const logoGrupo = "__LOGOG64__";
const RODAPE_INST = "__ENDERECO__";

const topo = document.getElementById('topo');
const busca = document.getElementById('busca');
const filtro = document.getElementById('filtro');
const contador = document.getElementById('contador');
const cards = [...document.querySelectorAll('.card')];
let lastQ = '';

function esc(s){ return String(s).replace(/[&<>"]/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c])); }
function norm(s){ return s.toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,''); }
function mapaNorm(s){ let r=''; for(const ch of s){ r += ch.normalize('NFD')[0].toLowerCase(); } return r; }
function aviso(t){ const a=document.getElementById('aviso'); a.textContent=t; a.style.display='block';
  clearTimeout(aviso._t); aviso._t=setTimeout(()=>a.style.display='none',1800); }

function setTopoH(){ document.documentElement.style.setProperty('--topo-h', topo.offsetHeight+'px'); }
window.addEventListener('resize', setTopoH);

/* ---------- filtros ---------- */
function chipsAtivos(g){ return [...document.querySelectorAll('.chip.on[data-g="'+g+'"]')].map(c=>c.dataset.v); }
function togglePainel(){ const p=document.getElementById('painel-filtros');
  p.style.display = p.style.display==='block' ? 'none':'block'; setTopoH(); }
function limpaChips(){ document.querySelectorAll('.chip.on').forEach(c=>c.classList.remove('on')); aplica(); }
document.querySelectorAll('.chip').forEach(c=>c.addEventListener('click',()=>{ c.classList.toggle('on'); aplica(); }));

function filtroAtivo(){ return !!(busca.value.trim() || filtro.value ||
  document.querySelector('.chip.on')); }

function aplica(){
  const q = norm(busca.value.trim());
  const st = filtro.value;
  const dOn = chipsAtivos('d'), fOn = chipsAtivos('f'), pOn = chipsAtivos('p');
  const ativo = filtroAtivo();
  let nLin=0, nSis=0;
  cards.forEach(card=>{
    if (st && card.id !== st){ card.classList.add('oculta'); return; }
    const tituloBate = q && card.dataset.titulo.includes(q);
    let vis=0;
    card.querySelectorAll('tbody tr').forEach(tr=>{
      let ok = !q || tituloBate || tr.dataset.busca.includes(q);
      if (ok && dOn.length) ok = tr.dataset.diam.split(' ').some(v=>dOn.includes(v));
      if (ok && fOn.length) ok = tr.dataset.fam.split('|').some(v=>fOn.includes(v));
      if (ok && pOn.length) ok = pOn.includes(tr.dataset.prot);
      tr.classList.toggle('oculta', !ok);
      if (ok) vis++;
    });
    const mostra = vis>0;
    card.classList.toggle('oculta', !mostra);
    if (mostra){ nSis++; nLin+=vis; if (ativo) card.classList.add('aberto'); }
  });
  document.getElementById('vazio').style.display = nSis ? 'none' : 'block';
  contador.textContent = nLin+' mangueiras em '+nSis+' sistemas';
  destaca(q);
}

function destaca(q){
  if (q===lastQ) {} 
  if (!q || q.length<2){
    if (lastQ) document.querySelectorAll('td .v').forEach(s=>{ s.textContent = s.dataset.o; });
    lastQ=q; return;
  }
  document.querySelectorAll('tbody tr').forEach(tr=>{
    const liga = !tr.classList.contains('oculta');
    tr.querySelectorAll('.v').forEach(s=>{
      const o = s.dataset.o;
      if (!liga){ if (s.firstElementChild) s.textContent=o; return; }
      const n = mapaNorm(o);
      let i=0, li=0, r='';
      while((i=n.indexOf(q,li))!==-1){ r+=esc(o.slice(li,i))+'<mark>'+esc(o.slice(i,i+q.length))+'</mark>'; li=i+q.length; }
      if (li===0){ s.textContent=o; } else { r+=esc(o.slice(li)); s.innerHTML=r; }
    });
  });
  lastQ=q;
}

function limpa(){
  busca.value=''; filtro.value='';
  document.querySelectorAll('.chip.on').forEach(c=>c.classList.remove('on'));
  cards.forEach(c=>c.classList.remove('aberto'));
  aplica();
}

/* ---------- cards / navegação ---------- */
function toggleCard(cab){ cab.closest('.card').classList.toggle('aberto'); }
function abreTodos(sim){ cards.forEach(c=>{ if(!c.classList.contains('oculta')) c.classList.toggle('aberto',sim); }); }
function irPara(st, rowId){
  const card=document.getElementById(st);
  if (!card) return;
  if (card.classList.contains('oculta')) limpa();
  card.classList.add('aberto');
  card.scrollIntoView({behavior:'smooth'});
  if (rowId){ const tr=document.querySelector('tr[data-id="'+rowId+'"]');
    if(tr){ tr.classList.remove('flash'); void tr.offsetWidth; tr.classList.add('flash'); } }
  if (history.replaceState) history.replaceState(null,'','#'+st);
}
window.addEventListener('hashchange',()=>{ const st=location.hash.slice(1); if(st) irPara(st); });
function copiaLink(e,st){ e.stopPropagation();
  copiar(location.href.split('#')[0]+'#'+st, 'Link do sistema copiado'); }

window.addEventListener('scroll',()=>{ document.getElementById('btn-topo').style.display =
  window.scrollY>600 ? 'block':'none'; });

/* ---------- ordenação ---------- */
function valorOrd(tr,k){
  if (k==='dia'){ const d=tr.dataset.diam.split(' ')[0]; return d?parseInt(d):9999; }
  const td=tr.children[parseInt(k)+2]; if(!td) return '';
  const sp=td.querySelector('.v'); const t=sp?sp.dataset.o:td.textContent.trim();
  const num=t.replace(',','.');
  return /^-?\d+(\.\d+)?$/.test(num) ? parseFloat(num) : norm(t);
}
document.querySelectorAll('th.th-sort').forEach(th=>th.addEventListener('click',()=>{
  const k=th.dataset.k, tabela=th.closest('table'), tb=tabela.querySelector('tbody');
  const dir = th.dataset.dir==='asc' ? 'desc':'asc';
  tabela.querySelectorAll('th').forEach(o=>{ o.dataset.dir=''; const d=o.querySelector('.dir'); if(d) d.textContent=''; });
  th.dataset.dir=dir; let f=th.querySelector('.dir');
  if(!f){ f=document.createElement('span'); f.className='dir'; th.appendChild(f); }
  f.textContent = dir==='asc'?' ▲':' ▼';
  const rows=[...tb.children];
  rows.sort((a,b)=>{ const va=valorOrd(a,k), vb=valorOrd(b,k);
    if (typeof va==='number' && typeof vb==='number') return dir==='asc'?va-vb:vb-va;
    return dir==='asc' ? String(va).localeCompare(String(vb)) : String(vb).localeCompare(String(va)); });
  rows.forEach(r=>tb.appendChild(r));
}));

/* ---------- copiar / onde é usada ---------- */
function copiar(txt,msg){
  const fim=()=>aviso(msg||'Copiado: '+txt);
  if (navigator.clipboard && navigator.clipboard.writeText){
    navigator.clipboard.writeText(txt).then(fim).catch(()=>copiaFallback(txt,fim));
  } else copiaFallback(txt,fim);
}
function copiaFallback(txt,fim){ const ta=document.createElement('textarea'); ta.value=txt;
  document.body.appendChild(ta); ta.select();
  try{ document.execCommand('copy'); fim(); }catch(e){ aviso('Não foi possível copiar'); }
  ta.remove(); }

const usoIdx=new Map();
document.querySelectorAll('tbody tr').forEach(tr=>{
  if (OEM_IDX<0) return;
  const td=tr.children[OEM_IDX+2]; if(!td) return;
  const sp=td.querySelector('.v'); if(!sp) return;
  const card=tr.closest('.card');
  sp.dataset.o.split(' / ').forEach(o=>{
    if(!usoIdx.has(o)) usoIdx.set(o,[]);
    usoIdx.get(o).push({st:card.id, nome:card.dataset.nome, id:tr.dataset.id});
  });
});
function abreUso(oem){
  const lista=usoIdx.get(oem)||[];
  document.getElementById('uso-titulo').textContent='OEM '+oem+' — usada em '+lista.length+' local(is)';
  const ul=document.getElementById('lista-uso'); ul.innerHTML='';
  lista.forEach(u=>{ const li=document.createElement('li');
    li.innerHTML=esc(u.nome)+' <span class="st">'+esc(u.st)+'</span>';
    li.addEventListener('click',()=>{ document.getElementById('modal-uso').classList.remove('aberto'); irPara(u.st,u.id); });
    ul.appendChild(li); });
  document.getElementById('modal-uso').classList.add('aberto');
}
document.querySelectorAll('.c-oem .v').forEach(sp=>sp.addEventListener('click',()=>abreUso(sp.dataset.o.split(' / ')[0])));
document.querySelectorAll('.ic-copia').forEach(b=>b.addEventListener('click',e=>{
  e.stopPropagation(); copiar(b.dataset.o,'OEM copiado: '+b.dataset.o); }));

/* ---------- cesta (requisição) ---------- */
const cesta=new Map();
function salvaCesta(){ try{ localStorage.setItem('cedro_cesta', JSON.stringify([...cesta])); }catch(e){} }
function badge(){ document.getElementById('badge-cesta').textContent=cesta.size; }
function toggleCesta(){ const c=document.getElementById('cesta');
  c.classList.toggle('aberta'); if (c.classList.contains('aberta')) montaCesta(); }
function dadosLinha(id){
  const tr=document.querySelector('tr[data-id="'+id+'"]'); if(!tr) return null;
  const card=tr.closest('.card');
  const vals=COLS.map((c,i)=>{ const sp=tr.children[i+2]?tr.children[i+2].querySelector('.v'):null;
    return sp?sp.dataset.o:''; });
  return {tr, vals, dia: tr.children[1].querySelector('.dnum') ? tr.children[1].querySelector('.dnum').textContent : '',
          nome: card.dataset.nome, st: card.id};
}
function montaCesta(){
  const corpo=document.getElementById('cesta-corpo');
  if(!cesta.size){ corpo.innerHTML='<p style="color:#6F7E63;font-size:.85rem">Marque as caixinhas das mangueiras para montar a lista.</p>'; return; }
  let h='<table><thead><tr><th>Qtd</th><th>OEM</th><th>Mang.</th><th>Compr.</th><th></th></tr></thead><tbody>';
  for (const [id,q] of cesta){
    const d=dadosLinha(id); if(!d) continue;
    const oem = OEM_IDX>=0 ? d.vals[OEM_IDX] : id;
    h+='<tr><td><input type="number" min="1" value="'+q+'" data-id="'+id+'"></td>'+
       '<td>'+esc(oem)+'</td><td>'+esc(d.vals[OEM_IDX>=0?OEM_IDX+1:0]||'')+'</td>'+
       '<td>'+esc(d.vals[OEM_IDX>=0?OEM_IDX+2:1]||'')+'</td>'+
       '<td><button class="rm" data-id="'+id+'" title="Remover">✕</button></td></tr>'+
       '<tr><td></td><td colspan="4" style="text-align:left;color:#6F7E63;font-size:.72rem;border-top:none">'+esc(d.nome)+'</td></tr>';
  }
  corpo.innerHTML=h+'</tbody></table>';
  corpo.querySelectorAll('input[type=number]').forEach(i=>i.addEventListener('change',()=>{
    cesta.set(i.dataset.id, Math.max(1,parseInt(i.value)||1)); salvaCesta(); }));
  corpo.querySelectorAll('.rm').forEach(b=>b.addEventListener('click',()=>{
    cesta.delete(b.dataset.id); const tr=document.querySelector('tr[data-id="'+b.dataset.id+'"]');
    if(tr) tr.querySelector('.sel').checked=false; salvaCesta(); badge(); montaCesta(); }));
}
function limpaCesta(){ cesta.forEach((q,id)=>{ const tr=document.querySelector('tr[data-id="'+id+'"]');
    if(tr) tr.querySelector('.sel').checked=false; });
  cesta.clear(); salvaCesta(); badge(); montaCesta(); }
document.querySelectorAll('input.sel').forEach(cb=>cb.addEventListener('change',()=>{
  const id=cb.closest('tr').dataset.id;
  if (cb.checked) cesta.set(id, cesta.get(id)||1); else cesta.delete(id);
  salvaCesta(); badge(); }));
document.querySelectorAll('input.sel-all').forEach(cb=>cb.addEventListener('click',e=>e.stopPropagation()));
document.querySelectorAll('input.sel-all').forEach(cb=>cb.addEventListener('change',()=>{
  cb.closest('table').querySelectorAll('tbody tr:not(.oculta) .sel').forEach(s=>{
    if (s.checked!==cb.checked){ s.checked=cb.checked; s.dispatchEvent(new Event('change')); } });
}));

/* ---------- PDFs / CSV ---------- */
function descricaoFiltro(){
  const p=[];
  if (busca.value.trim()) p.push('busca: “'+busca.value.trim()+'”');
  if (filtro.value){ const c=document.getElementById(filtro.value);
    p.push('sistema: '+(c?c.dataset.nome:filtro.value)); }
  const chips=[...document.querySelectorAll('.chip.on')].map(c=>c.textContent.trim());
  if (chips.length) p.push('filtros: '+chips.join(', '));
  return p.length ? p.join(' · ') : 'sem filtros (catálogo completo)';
}
function capa(titulo, docnum){
  return '<div class="rel-topo">'+
    (logoCedro?'<img class="rel-logo" src="'+logoCedro+'">':'')+
    '<div class="rel-tit"><h1>'+titulo+'</h1><div class="linha2">Usina Cedro · Pedra Agroindustrial · Catálogo Copecar · emitido em '+
    new Date().toLocaleString('pt-BR')+'</div></div>'+
    (docnum?'<div class="rel-doc">'+docnum+'</div>':'')+
    (logoGrupo?'<img class="rel-logo dir" src="'+logoGrupo+'">':'')+
  '</div><div class="rel-fio"></div>';
}
function thsRelatorio(){ return '<th style="width:7%">Ø</th>'+COLS.map(c=>'<th>'+esc(c)+'</th>').join(''); }
function abrePdfOpcoes(){
  document.getElementById('pdf-resumo').textContent='Seleção atual — '+contador.textContent+' ('+descricaoFiltro()+').';
  document.getElementById('pdf-opcoes').classList.add('aberto');
}
function fechaPdfOpcoes(){ document.getElementById('pdf-opcoes').classList.remove('aberto'); }
function imprime(html){
  const rel=document.getElementById('relatorio'); rel.innerHTML=html;
  document.body.classList.add('imprimindo'); window.print();
}
window.addEventListener('afterprint',()=>{ document.body.classList.remove('imprimindo');
  document.getElementById('relatorio').innerHTML=''; });

function gerarPDF(){
  const incluirImg=document.getElementById('chk-img').checked;
  let h=capa('Relatório de Mangueiras — Colhedora John Deere CH570')+
    '<div class="rel-filtros"><b>Filtros aplicados:</b> '+esc(descricaoFiltro())+' &nbsp;·&nbsp; <b>'+contador.textContent+'</b></div>';
  cards.forEach(card=>{
    if (card.classList.contains('oculta')) return;
    const linhas=[...card.querySelectorAll('tbody tr')].filter(t=>!t.classList.contains('oculta'));
    if (!linhas.length) return;
    h+='<div class="rel-sis"><h3>'+esc(card.dataset.nome)+' — '+esc(card.id)+'</h3>';
    if (incluirImg){ const img=card.querySelector('img.diagrama'); if(img) h+='<img src="'+img.src+'">'; }
    h+='<table><thead><tr>'+thsRelatorio()+'</tr></thead><tbody>';
    linhas.forEach(tr=>{
      h+='<tr><td>'+tr.children[1].innerHTML+'</td>';
      COLS.forEach((c,i)=>{ const sp=tr.children[i+2]?tr.children[i+2].querySelector('.v'):null;
        h+='<td>'+esc(sp?sp.dataset.o:'')+'</td>'; });
      h+='</tr>';
    });
    h+='</tbody></table></div>';
  });
  h+='<div class="rel-rodape">Catálogo de Mangueiras CH570 · '+RODAPE_INST+'</div>';
  fechaPdfOpcoes(); imprime(h);
}

function gerarRequisicao(){
  if (!cesta.size){ aviso('A lista de requisição está vazia'); return; }
  const ag=new Date(), p2=v=>String(v).padStart(2,'0');
  const docnum='REQ-'+ag.getFullYear()+p2(ag.getMonth()+1)+p2(ag.getDate())+'-'+p2(ag.getHours())+p2(ag.getMinutes());
  let totalQtd=0; cesta.forEach(q=>totalQtd+=q);
  let h=capa('Requisição de Peças — Mangueiras CH570', docnum)+
    '<div class="req-meta">'+
      '<div><b>Documento</b>'+docnum+'</div>'+
      '<div><b>Itens distintos</b>'+cesta.size+'</div>'+
      '<div><b>Total de unidades</b>'+totalQtd+'</div>'+
    '</div>'+
    '<table><thead><tr><th style="width:5%">Nº</th><th style="width:7%">Qtd</th><th style="width:6%">Ø</th>'+
    COLS.map(c=>'<th>'+esc(c)+'</th>').join('')+'<th>Sistema</th></tr></thead><tbody>';
  let n=0;
  for (const [id,q] of cesta){
    const d=dadosLinha(id); if(!d) continue; n++;
    h+='<tr><td>'+n+'</td><td class="qtd">'+q+'</td><td>'+d.tr.children[1].innerHTML+'</td>'+
       d.vals.map(v=>'<td>'+esc(v)+'</td>').join('')+
       '<td style="font-size:7.5pt">'+esc(d.nome)+' ('+esc(d.st)+')</td></tr>';
  }
  h+='</tbody></table>'+
     '<div class="req-total">Total: <b>'+cesta.size+' item(ns) · '+totalQtd+' unidade(s)</b></div>'+
     '<div class="req-obs"><b>Observações</b><div class="linha"></div><div class="linha"></div></div>'+
     '<div class="req-assin"><h4>Assinaturas e autorização</h4><div class="campos">'+
       '<div class="campo"><div class="linha-ass"></div><b>Solicitante</b><br><span>Nome legível / Data</span></div>'+
       '<div class="campo"><div class="linha-ass"></div><b>Autorizado por</b><br><span>Nome legível / Data</span></div>'+
     '</div></div>'+
     '<div class="rel-rodape">'+RODAPE_INST+' · '+docnum+'</div>';
  toggleCesta(); imprime(h);
}

function exportaCSV(){
  if (!cesta.size){ aviso('A lista de requisição está vazia'); return; }
  const sep=';';
  let csv='\uFEFF'+['QTD','DIAM',...FULLCOLS,'SISTEMA','ST'].join(sep)+'\r\n';
  for (const [id,q] of cesta){
    const d=dadosLinha(id); if(!d) continue;
    const cel=v=>'"'+String(v).replace(/"/g,'""')+'"';
    csv+=[q, d.dia, ...d.vals.map(cel), cel(d.nome), d.st].join(sep)+'\r\n';
  }
  const blob=new Blob([csv],{type:'text/csv;charset=utf-8'});
  const a=document.createElement('a'); a.href=URL.createObjectURL(blob);
  a.download='requisicao_mangueiras_CH570.csv'; a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),4000);
}

/* ---------- modal de zoom (rolar = zoom, arrastar = mover, pinça) ---------- */
const modal=document.getElementById('modal'), mimg=document.getElementById('modal-img');
let zs=1, ztx=0, zty=0;
const aplicaZ=()=>{ mimg.style.transform='translate('+ztx+'px,'+zty+'px) scale('+zs+')'; };
function abreModal(src){
  mimg.src=src; modal.classList.add('aberto');
  mimg.onload=()=>{ const r=Math.min((window.innerWidth-40)/mimg.naturalWidth,
                                     (window.innerHeight-40)/mimg.naturalHeight, 1);
    zs=r; ztx=(window.innerWidth-mimg.naturalWidth*r)/2; zty=(window.innerHeight-mimg.naturalHeight*r)/2; aplicaZ(); };
  if (mimg.complete && mimg.naturalWidth) mimg.onload();
}
function fechaModal(){ modal.classList.remove('aberto'); }
function zoomEm(f,cx,cy){ const novo=Math.min(Math.max(zs*f,0.2),10);
  ztx = cx-(cx-ztx)*(novo/zs); zty = cy-(cy-zty)*(novo/zs); zs=novo; aplicaZ(); }
function zoomBtn(f){ zoomEm(f, window.innerWidth/2, window.innerHeight/2); }
function zoomReset(){ if (mimg.onload) mimg.onload(); }
modal.addEventListener('wheel',e=>{ e.preventDefault(); zoomEm(e.deltaY<0?1.18:0.85,e.clientX,e.clientY); },{passive:false});
modal.addEventListener('dblclick',e=>{ if(e.target===mimg) zoomEm(zs<1.5?2.2:0.3,e.clientX,e.clientY); });
modal.addEventListener('click',e=>{ if(e.target===modal) fechaModal(); });
const dedos=new Map(); let distAnt=0;
modal.addEventListener('pointerdown',e=>{ dedos.set(e.pointerId,[e.clientX,e.clientY]);
  modal.setPointerCapture(e.pointerId); mimg.style.cursor='grabbing'; });
modal.addEventListener('pointermove',e=>{
  if (!dedos.has(e.pointerId)) return;
  if (dedos.size===1){ const [px,py]=dedos.get(e.pointerId);
    ztx+=e.clientX-px; zty+=e.clientY-py; aplicaZ(); }
  dedos.set(e.pointerId,[e.clientX,e.clientY]);
  if (dedos.size===2){ const pts=[...dedos.values()];
    const dist=Math.hypot(pts[0][0]-pts[1][0],pts[0][1]-pts[1][1]);
    if (distAnt) zoomEm(dist/distAnt,(pts[0][0]+pts[1][0])/2,(pts[0][1]+pts[1][1])/2);
    distAnt=dist; }
});
['pointerup','pointercancel'].forEach(ev=>modal.addEventListener(ev,e=>{
  dedos.delete(e.pointerId); distAnt=0; mimg.style.cursor='grab'; }));
document.querySelectorAll('img.diagrama').forEach(img=>img.addEventListener('click',()=>abreModal(img.src)));

/* ---------- teclado ---------- */
document.addEventListener('keydown',e=>{
  if (e.key==='Escape'){
    if (modal.classList.contains('aberto')) return fechaModal();
    const uso=document.getElementById('modal-uso'); if (uso.classList.contains('aberto')) return uso.classList.remove('aberto');
    if (document.getElementById('pdf-opcoes').classList.contains('aberto')) return fechaPdfOpcoes();
    const c=document.getElementById('cesta'); if (c.classList.contains('aberta')) return toggleCesta();
    if (filtroAtivo()) limpa();
    return;
  }
  if (e.key==='/' && !['INPUT','SELECT','TEXTAREA'].includes(document.activeElement.tagName)){
    e.preventDefault(); busca.focus();
  }
});

/* ---------- inicialização ---------- */
busca.addEventListener('input', aplica);
filtro.addEventListener('change', ()=>{ aplica(); if (filtro.value) irPara(filtro.value); });
try{ const j=localStorage.getItem('cedro_cesta');
  if (j) for (const [id,q] of JSON.parse(j)){
    const tr=document.querySelector('tr[data-id="'+id+'"]');
    if (tr){ cesta.set(id,q); tr.querySelector('.sel').checked=true; } }
}catch(e){}
badge(); setTopoH(); aplica();
if (location.hash.length>1) setTimeout(()=>irPara(location.hash.slice(1)),150);
</script>
</body>
</html>"""

# ----------------------------- MONTAGEM --------------------------------------
def gera_html(ordem, sistemas, colunas, dir_img):
    total = sum(len(v) for v in sistemas.values())
    legenda = dir_img / "legenda.jpg"
    legenda_b64 = b64(legenda) if legenda.exists() else ""
    def acha_logo(cands):
        for cand in cands:
            p = Path(cand)
            if p.exists():
                print(f"  logo embutido: {cand}")
                return b64(p)
        return ""
    logo_cedro = acha_logo(LOGOS_CEDRO)
    logo_grupo = acha_logo(LOGOS_GRUPO)

    oem_idx = colunas.index("OEM") if "OEM" in colunas else -1
    todos_d, todos_f, todos_p = set(), set(), set()
    opcoes, cards = [], []
    rid = 0

    for nome, st in ordem:
        regs = sistemas[(nome, st)]
        opcoes.append(f'<option value="{html.escape(st)}">{html.escape(nome)}</option>')
        imgs = (sorted(dir_img.glob(f"{st}.jpg")) + sorted(dir_img.glob(f"{st}_*.jpg"))) if st else []
        tag_imgs = "".join(f'<img class="diagrama" loading="lazy" src="{b64(p)}" '
                           f'alt="Diagrama {html.escape(st)}">' for p in imgs)
        ths = ('<th class="c-sel"><input type="checkbox" class="sel-all" '
               'title="Selecionar visíveis deste sistema"></th>'
               '<th class="c-dia th-sort" data-k="dia" title="Diâmetro">Ø</th>')
        for i, c in enumerate(colunas):
            ths += (f'<th class="th-sort" data-k="{i}" title="{html.escape(c)} — clique para ordenar">'
                    f'{html.escape(ABREVIACOES.get(c, c))}</th>')
        trs = []
        for r in regs:
            rid += 1
            mang = r.get(COL_MANGUEIRA, "")
            diams, fams = parse_mangueira(mang)
            prot = r.get(COL_PROTECAO, "")
            todos_d |= diams; todos_f |= fams
            if prot: todos_p.add(prot)
            busca_txt = normaliza(" ".join([nome, st] + [r.get(c, "") for c in colunas]))
            pontos = "".join(ponto_diam(d) for d in sorted(diams))
            dnum = "".join(f'<span class="dnum">-{d}</span>' for d in sorted(diams))
            tds = (f'<td class="c-sel"><input type="checkbox" class="sel" '
                   f'aria-label="Selecionar"></td><td class="c-dia">{pontos}{dnum}</td>')
            for c in colunas:
                v = html.escape(r.get(c, ""))
                if c == "OEM":
                    tds += (f'<td class="c-oem"><span class="v" data-o="{v}" '
                            f'title="Ver onde esta peça é usada">{v}</span>'
                            f'<button class="ic-copia" data-o="{v}" title="Copiar OEM">⧉</button></td>')
                else:
                    tds += f'<td><span class="v" data-o="{v}">{v}</span></td>'
            trs.append(f'<tr data-id="r{rid}" data-busca="{html.escape(busca_txt)}" '
                       f'data-diam="{" ".join(str(d) for d in sorted(diams))}" '
                       f'data-fam="{html.escape("|".join(sorted(fams)))}" '
                       f'data-prot="{html.escape(prot)}">{tds}</tr>')
        cards.append(f"""
<section class="card" id="{html.escape(st)}" data-nome="{html.escape(nome)}"
         data-titulo="{html.escape(normaliza(nome + ' ' + st))}">
  <header class="card-cab" onclick="toggleCard(this)">
    <span class="seta">▶</span>
    <h2>{html.escape(nome)}</h2>
    <span class="badge">{html.escape(st)}</span>
    <span class="qtd">{len(regs)} mangueira(s)</span>
    <button class="ic-link" title="Copiar link direto deste sistema"
            onclick="copiaLink(event,'{html.escape(st)}')">🔗</button>
  </header>
  <div class="card-corpo">
    <div class="bloco-img">{tag_imgs or '<p class="sem-img">Sem diagrama</p>'}</div>
    <table><thead><tr>{ths}</tr></thead><tbody>{''.join(trs)}</tbody></table>
  </div>
</section>""")

    chips_d = "".join(
        f'<button class="chip" data-g="d" data-v="{d}">{ponto_diam(d)} -{d}</button>'
        for d in sorted(todos_d))
    chips_f = "".join(
        f'<button class="chip" data-g="f" data-v="{html.escape(f)}">{html.escape(f)}</button>'
        for f in sorted(todos_f))
    chips_p = "".join(
        f'<button class="chip" data-g="p" data-v="{html.escape(p)}">'
        f'{html.escape("— sem" if p == "-" else p)}</button>'
        for p in sorted(todos_p))

    btn_leg = ('<button onclick="abreModal(legendaSrc)">Legenda de cores</button>'
               if legenda_b64 else "")
    logo_tag = f'<img class="logo" src="{logo_cedro}" alt="Usina Cedro">' if logo_cedro else ""
    logo_tag_g = (f'<img class="logo logo-dir" src="{logo_grupo}" alt="Pedra Agroindustrial">'
                  if logo_grupo else "")

    return (TEMPLATE
            .replace("__NSIS__", str(len(ordem)))
            .replace("__NLIN__", str(total))
            .replace("__DATA__", datetime.now().strftime("%d/%m/%Y %H:%M"))
            .replace("__LOGOG64__", logo_grupo)
            .replace("__LOGOC__", logo_cedro)
            .replace("__LOGOG__", logo_tag_g)
            .replace("__LOGO__", logo_tag)
            .replace("__ENDERECO__", html.escape(ENDERECO_UNIDADE))
            .replace("__OPCOES__", "".join(opcoes))
            .replace("__CHIPS_D__", chips_d)
            .replace("__CHIPS_F__", chips_f)
            .replace("__CHIPS_P__", chips_p)
            .replace("__CARDS__", "".join(cards))
            .replace("__BTNLEG__", btn_leg)
            .replace("__LEG__", legenda_b64)
            .replace("__COLSJS__", json.dumps([ABREVIACOES.get(c, c) for c in colunas],
                                              ensure_ascii=False))
            .replace("__FULLCOLSJS__", json.dumps(colunas, ensure_ascii=False))
            .replace("__OEMIDX__", str(oem_idx)))

# ----------------------------- MAIN ------------------------------------------
if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Gera o catálogo HTML interativo a partir do Excel.")
    ap.add_argument("--excel", default="Catalogo_CH570.xlsx")
    ap.add_argument("--diagramas", default="diagramas")
    ap.add_argument("--saida", default="Catalogo_CH570.html")
    a = ap.parse_args()

    dir_img = Path(a.diagramas)
    if not Path(a.excel).exists():
        sys.exit(f"Excel não encontrado: {a.excel}")
    if not dir_img.is_dir():
        print(f"Aviso: pasta '{a.diagramas}' não encontrada — catálogo sairá sem imagens.")

    ordem, sistemas, colunas = le_excel(a.excel)
    htm = gera_html(ordem, sistemas, colunas, dir_img)
    Path(a.saida).write_text(htm, encoding="utf-8")
    tam = Path(a.saida).stat().st_size / 1e6
    print(f"✔ {a.saida} gerado: {len(ordem)} sistemas, "
          f"{sum(len(v) for v in sistemas.values())} mangueiras, "
          f"colunas: {', '.join(colunas)} ({tam:.1f} MB)")
